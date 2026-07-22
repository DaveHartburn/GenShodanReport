#!/usr/bin/python

# Generate a shodan summary report, showing various aspects of exposure
# By Dave Hartburn - March 2023

# Note - This works on the download data and processes the CSV. For more advanced stuff it should
# probably use the API

# Usage:
# genShodanReport.py [-q] [-l filename] [-o output directory] -s <search>
#   -q : Quiet, limited screen output
#   -l : Followed by a filename, use a previously downloaded file (useful for testing)
#         The download zipped JSON file (dlData.json.gz) is the one required
#   -o : Specify the output directory, otherwise one will be made with the date and time. Data will be overwritten
#   search : A search string. e.g. net:1.1.0.0/16, ignored if we are using a local file
#            Surround in quotes for a more complicated search, e.g. with excludes "net:1.1.0.0/16 -net:1.1.225.0/24"
#   Note, shodan commands will not execute if you have not configured your API string
import os, sys, getopt, time, subprocess
import gzip, json
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from pprint import pprint

# Set some defaults
dlFile="dlData"		# Default name for downloaded data (.json.gz added automatically)
outputFile="report.txt"
searchStr = ""
quiet=False
localFile=""
outputDir=""
makeExcel=True

# Parse command line arguments
argList=sys.argv[1:]
opts="ql:s:o:X"

try:
	args, vals = getopt.getopt(argList, opts)

	for curArg, curVal in args:
		if curArg=="-q":
			quiet=True
		if curArg=="-l":
			localFile=curVal
		if curArg=="-s":
			searchStr=curVal
            #print("Search string=",searchStr)
		if curArg=="-o":
			outputDir=curVal
		if curArg=="-X":
			makeExcel=False

except getopt.error as err:
	print(str(err))
#exit(0)
if(searchStr=="" and localFile==""):
	print("Error: You must specify either a search string or a local download filename")
	print("Usage:")
	print("genShodanReport.py [-q] [-l filename] -s <search>")
	print("  -q : Quiet, limited screen output")
	print("  -l : Followed by a filename, use a previously downloaded file (useful for testing)")
	print("  search : A search string. e.g. net:1.1.0.0/16, ignored if we are using a local file")
	print("  -X : Do not produce an Excel spreadhsheet")
	print("Note, shodan commands will not execute if you have not configured your API string")
	exit(1)

# ********************************************
def myPrint(msg):
	if(quiet==False):
		print(msg)

def writeToFile(fh, msg):
	# Write to file and possibly output to screen
	if(quiet==False):
		print(msg)
	# To file
	fh.write(msg+"\n")

# *********************************************

myPrint("Going ahead with quiet={} localfile={} search={}".format(quiet, localFile, searchStr))
# Generate a local time format for storing data

timeString=time.strftime("%Y_%m_%d_%H%M")
if outputDir=="":
	outputDir="shodan_"+timeString

# *** Warning, need to watch for OS independence here ***
myPrint("Creating local directory for results, "+outputDir)
os.mkdir(outputDir)
# Filename for excel file
excelFile=os.path.join(outputDir, "attackSurface_"+timeString+".xlsx")
#print("Producing excel file ", excelFile)


# localFile will become the json.gz input
if(localFile==""):
	# Not reading from a local file, download
	myPrint("Downloading data...")
	print(outputDir, dlFile)
	cmd=["shodan", "download", "--limit", "-1", os.path.join(outputDir, dlFile), searchStr]
	subprocess.run(cmd)
	print(cmd)
	localFile=os.path.join(outputDir, dlFile+".json.gz")

#exit(1)
# Convert to CSV
cmd=["shodan", "convert", "--fields", "ip,ip_str,hostnames,transport,port,ssl.cipher.versions", localFile, "csv"]
myPrint("Creating CSV file...")
subprocess.run(cmd)
# Annoyingly this writes to the same directory as the input file.
#csvFile=dlFile+."csv"

# Set up data structures for reporting
openIPs={}
openPorts={}
sslVersions={}
subnets={}

# Open an output file for writing
ofile=open(os.path.join(outputDir, outputFile), "w")

# Read the download file, a line at a time. Each line is it's own JSON
fin = gzip.open(localFile, 'r')
for line in fin:
	line = line.decode('utf-8')
	jdata = json.loads(line)
	#pprint(jdata)

	ip=jdata["ip_str"]
	trans=jdata["transport"]
	port=str(jdata["port"])
	proto=trans+"/"+port

	# Have we recorded this IP? If not create
	if ip in openIPs:
		# Yes, add to port list and increase port count by one
		openIPs[ip]["portCount"]+=1
		openIPs[ip]["portList"].append(proto)
	else:
		# No, create
		openIPs[ip]={}
		openIPs[ip]["portCount"]=1
		openIPs[ip]["portList"]=[proto]
		openIPs[ip]["hostnames"]=jdata["hostnames"]

	# Count open ports
	if proto in openPorts:
		# Seen this before
		openPorts[proto]+=1
	else:
		openPorts[proto]=1

	# Count SSL versions if data present
	if "ssl" in jdata:
		if "versions" in jdata["ssl"]:
			#pprint(jdata["ssl"]["versions"])
			for ver in jdata["ssl"]["versions"]:
				if(ver[0]!="-"):
					if ver in sslVersions:
						sslVersions[ver]+=1
					else:
						sslVersions[ver]=1

	try:
		# Look at subnets
		ipArr=ip.split('.')
		sub=ipArr[2]
		#myPrint("  This IP is in subnet {}".format(ipArr[2]))
		if sub in  subnets:
			subnets[sub]+=1
		else:
			subnets[sub]=1
	except IndexError:
		# Error processing - this is probably IPv6 and I need to do something more sensible about this
		print("Warning: unable to split subnet information into subnets for subnet count. Ignoring: ", ipArr)
#pprint(subnets)
# Test output
#myPrint("openIPs data structure:")
#pprint(openIPs)
#myPrint("Protocol count")
#pprint(openPorts)
#myPrint("SSL version count")
#pprint(sslVersions)



# Produce summary report
writeToFile(ofile, "Number of hosts visible to the internet={}".format(len(openIPs)))
writeToFile(ofile, "Number of different protocols open={}".format(len(openPorts)))
writeToFile(ofile, "Open IP addresses:")
writeToFile(ofile, "IP,Hostnames,Num ports,Port List")
# Difficult to sort into order, but this is to be pasted in Excel anyway
for i in openIPs:
	#pprint(openIPs[i])
	hnames=" ".join(openIPs[i]["hostnames"])
	ports=" ".join(openIPs[i]["portList"])
	writeToFile(ofile, "{},{},{},{}".format(i,hnames, openIPs[i]["portCount"],ports))
sortPorts=sorted(openPorts.items(), key=lambda x:x[1], reverse=True)
#print(sortPorts)
writeToFile(ofile, "\n\nOpen ports:")
writeToFile(ofile, "port, count")
for p, v in sortPorts:
	writeToFile(ofile, "{},{}".format(p, v))

sortSSL=sorted(sslVersions.items(), key=lambda x:x[1], reverse=True)
writeToFile(ofile, "\n\nSSL versions in use:")
writeToFile(ofile, "Version, count")
for s,v  in sortSSL:
	writeToFile(ofile, "{},{}".format(s, v))

sortSubnets=sorted(subnets.items(), key=lambda x:x[1], reverse=True)
writeToFile(ofile, "\n\nSubnet distribution:")
for s, v in sortSubnets:
	writeToFile(ofile, "{},{}".format(s,v))
ofile.close()

# If in quite mode, only output a summary
if(quiet):
	print("Number of hosts visible to the internet={}".format(len(openIPs)))
	print("Number of different protocols open={}".format(len(openPorts)))

# Excel output?
# Produce Excel spreasheet
if makeExcel:
	workbook=Workbook()
	# Create the sheets, 'sheet' already created by default
	sheet=workbook.active
	sheet.title="Summary"
	workbook.create_sheet("Condensed")
	workbook.create_sheet("Per Proto")

	# Write the summary tab
	sheet = workbook["Summary"]
	sheet["A1"] = "Exposed IPs"
	sheet["A1"].font = Font(bold=True)
	sheet["B1"] = len(openIPs)
	sheet["A2"] = "Exposed Protocols"
	sheet["A2"].font = Font(bold=True)
	sheet["B2"] = len(openPorts)

	sheet["A4"]= "Protocol"
	sheet["B4"] = "Count"

	l=5	# Line where data starts
	tabRef="A4:B"
	sortPorts=sorted(openPorts.items(), key=lambda x:x[1], reverse=True)
	for p, v in sortPorts:
		#print(p,v)
		sheet.cell(row=l, column=1).value=p
		sheet.cell(row=l, column=2).value=v
		l+=1
	tabRef+=str(l-1)
	#print(tabRef)
	table = Table(displayName="ProtoCount", ref=tabRef)
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
							showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	table.tableStyleInfo = style
	sheet.add_table(table)
	# Set width and lead summery style
	sheet.column_dimensions["A"].width = 17

	# Add the condensed sheet (one entry per IP)
	sheet=workbook["Condensed"]
	# Create the per proto at the same time
	prosheet=workbook["Per Proto"]

	# Create table headers
	sheet.append(["IP", "Hostnames", "Num ports", "Port List"])
	prosheet.append(["IP", "Hostnames", "Num ports", "Protocol"])
	# Line number trackers
	sheetl=1
	prol=1

	for i in openIPs:
		hnames=" ".join(openIPs[i]["hostnames"])
		ports=" ".join(openIPs[i]["portList"])
		#print(openIPs[i]["portList"])

		# Add to condensed sheet
		sheet.append([i,hnames, openIPs[i]["portCount"],ports])
		sheetl+=1

		# Add to per proto sheet
		for p in openIPs[i]["portList"]:
			prosheet.append([i, hnames, openIPs[i]["portCount"], p])
			prol+=1

	# Turn data into tables
	condRef="A1:D"+str(sheetl)
	proRef="A1:D"+str(prol)
	ctable = Table(displayName="CondensedProto", ref=condRef)
	ptable = Table(displayName="Protocols", ref=proRef)
	# Create styles
	cstyle = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=False,
							showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	pstyle = TableStyleInfo(name="TableStyleMedium11", showFirstColumn=False,
							showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	ctable.tableStyleInfo=cstyle
	ptable.tableStyleInfo=pstyle
	sheet.add_table(ctable)
	prosheet.add_table(ptable)

	# Set the column widths
	cwid = {"A":18, "B":50, "C":11, "D":35}
	pwid = {"A":18, "B":50, "C":11, "D":15}
	for c in cwid:
		sheet.column_dimensions[c].width = cwid[c]
	for c in pwid:
		prosheet.column_dimensions[c].width = pwid[c]
	workbook.save(filename=excelFile)